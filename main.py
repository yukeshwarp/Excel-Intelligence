import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
from openai import AzureOpenAI
import os
import openpyxl
import base64
from io import BytesIO
import xlwings as xw
from PIL import Image
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as openpyxlImage

st.title("Excel Intelligence")
client = AzureOpenAI(
    azure_endpoint=os.getenv("ENDPOINT"),
    api_key=os.getenv("OPENAI_API_KEY"),
    api_version="2024-10-01-preview",
)


def detect_graphs(file_path):
    try:
        wb = openpyxl.load_workbook(file_path)
        for sheet in wb.worksheets:
            if sheet._charts:  # Checking for charts in the worksheet
                return True
    except Exception as e:
        print(f"Error processing file: {e}")
        return False

    return False


# def encode_file_to_base64(file):
#     file_content = file.read()
#     # Reset the file pointer after reading
#     file.seek(0)
#     return base64.b64encode(file_content).decode("utf-8")


def extract_and_encode_charts(uploaded_file):
    if uploaded_file is None:
        return []
    try:
        # Read Excel file using openpyxl to access both data and charts
        wb = load_workbook(uploaded_file, keep_links=False)
        sheet = wb.active

        # Initialize a list to store all images in the sheet (graphs, etc.)
        images = []

        # Extract any images (graphs, charts) from the sheet
        # for image in sheet._image:
        #     img = openpyxlImage(image.ref)
        #     img_bytes = BytesIO(img.image)
        #     images.append(img_bytes)

        # Read the tabular data
        df = pd.read_excel(uploaded_file, sheet_name=0)

        # Convert dataframe to image using matplotlib
        fig, ax = plt.subplots(
            figsize=(10, len(df) * 0.5)
        )  # Adjust height based on rows
        ax.axis("tight")
        ax.axis("off")
        ax.table(
            cellText=df.values, colLabels=df.columns, cellLoc="center", loc="center"
        )

        # Save table as image (this will be the main image for the table)
        table_buf = BytesIO()
        plt.savefig(table_buf, format="png", bbox_inches="tight", dpi=300)
        table_buf.seek(0)

        # Create a base64 image for the table
        table_img = Image.open(table_buf)
        buffered = BytesIO()
        table_img.save(buffered, format="PNG")
        table_base64 = base64.b64encode(buffered.getvalue()).decode("utf-8")

        # Prepare to combine the table image and any charts (if found)
        base64_images = [table_base64]

        for img in images:
            img.seek(0)
            img_data = base64.b64encode(img.read()).decode("utf-8")
            base64_images.append(img_data)

        return base64_images

    except Exception as e:
        st.error(e)


with st.sidebar:
    uploaded_file = st.file_uploader("Upload Financials CSV", type=["xlsx"])
    if uploaded_file is not None:
        df = pd.read_excel(uploaded_file)

        if st.button("View Cleaned Data Table"):
            st.write("### Cleaned Data Table")
            st.dataframe(df)

if "messages" not in st.session_state:
    st.session_state.messages = []

if uploaded_file is not None:
    string_data = df.to_string()
    numerical_summary = df.describe().to_string()

    if "initial_analysis" not in st.session_state:
        analysis_prompt = f"""INSTRUCTION:
        Perform a comprehensive analysis of the provided financial dataset. Summarize key insights, trends, and any potential anomalies.
        Ensure the response is clear and easy to understand.
        
        KEY METRICS:
        {numerical_summary}
        
        INPUT: {string_data}"""

        with st.spinner("Analyzing spreadsheet..."):
            analysis_response = client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {
                        "role": "system",
                        "content": "You are an expert financial data analyst capable of extracting meaningful insights from spreadsheet data.",
                    },
                    {"role": "user", "content": analysis_prompt},
                ],
                temperature=0.0,
            )

        st.session_state.initial_analysis = analysis_response.choices[
            0
        ].message.content.strip()
        st.session_state.messages.append(
            {"role": "assistant", "content": st.session_state.initial_analysis}
        )

        persona_prompt = f"""
        Who can analyse this spreadsheet content with atmost expertise.
        Describe about their persona, tone and professional qualification if any.
        KEY METRICS:
        {numerical_summary}
        
        INPUT: {string_data}"""

        with st.spinner("Analyzing spreadsheet..."):
            analysis_response = client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {
                        "role": "system",
                        "content": "You are an expert in analysing data description.",
                    },
                    {"role": "user", "content": persona_prompt},
                ],
                temperature=0.0,
            )

        st.session_state.persona_analysis = analysis_response.choices[
            0
        ].message.content.strip()
        st.session_state.messages.append(
            {"role": "assistant", "content": st.session_state.persona_analysis}
        )
        base64_encoded_sheet = extract_and_encode_charts(uploaded_file)

        if len(base64_encoded_sheet) > 0:
            with st.spinner("Analyzing spreadsheet..."):
                analysis_response = client.chat.completions.create(
                    model="gpt-4o",
                    messages=[
                        {
                            "role": "system",
                            "content": "You are an expert financial data analyst capable of extracting meaningful insights from spreadsheet data. You explain the visual respresentations like graphs from the given base64 encoded spreadsheet.",
                        },
                        {
                            "role": "user",
                            "content": [
                                {
                                    "type": "text",
                                    "text": "Explain about the graphs of financial data you observe in the provided image. Explain about only the charts given in the image, don't add anyother explaination.",
                                },
                                {
                                    "type": "image_url",
                                    "image_url": {
                                        "url": f"data:image/png;base64,{base64_encoded_sheet[0]}"
                                    },
                                },
                            ],
                        },
                    ],
                    temperature=0.0,
                )

            st.session_state.image_analysis = analysis_response.choices[
                0
            ].message.content.strip()
            st.session_state.messages.append(
                {
                    "role": "assistant",
                    "content": f"Image analysis:\n{st.session_state.image_analysis}",
                }
            )
        else:
            st.error("Not received the image explaination.")

    for message in st.session_state.messages:
        with st.chat_message(message["role"]):
            st.markdown(message["content"])

    user_input = st.chat_input("Ask a question about the table:")

    if user_input:
        st.session_state.messages.append({"role": "user", "content": user_input})
        with st.chat_message("user"):
            st.markdown(user_input)

        prompt = f"""INSTRUCTION:
        You are a highly skilled AI assistant specializing in spreadsheet data analysis. Your task is to extract accurate answers from structured table data based on user queries.
        The provided input consists of tabular data formatted as a string, where each cell is represented by a pair consisting of a cell address and its content, separated by a comma (e.g., 'A1,Year').
        Cells are organized in a row-major order and are separated by '|', like 'A1,Year|A2,Profit'. Some cells may have empty values, represented as 'A1, |A2,Profit'.
        
        Your response should:
        - Accurately interpret the data in the table.
        - Provide numerical summaries with **consistent totals, averages, min, max, and standard deviations**.
        - Identify relevant patterns, trends, and numerical insights where applicable.
        - Ensure all calculations reference the precomputed metrics.
        
        KEY METRICS:
        {numerical_summary}
        
        INPUT: {string_data}
        USER QUESTION: {user_input}"""
        print(prompt)
        with st.spinner("Fetching answer..."):
            response_stream = client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {
                        "role": "system",
                        "content": f"""You are an advanced data analysis assistant, specialized in interpreting and extracting insights from structured spreadsheet data. 
                        You can accurately understand tabular information, identify key trends, and provide concise, human-readable responses based on the given financial dataset.
                        You possess the following characteristics:
                        {st.session_state.persona_analysis}
                        """,
                    },
                    {"role": "assistant", "content": st.session_state.initial_analysis},
                    {"role": "user", "content": prompt},
                ],
                temperature=0.0,
                stream=True,
            )

        bot_response = ""
        with st.chat_message("assistant"):
            response_placeholder = st.empty()
            for chunk in response_stream:
                if chunk.choices:
                    bot_response += chunk.choices[0].delta.content or ""
                    response_placeholder.markdown(bot_response)

        st.session_state.messages.append({"role": "assistant", "content": bot_response})
